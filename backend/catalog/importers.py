from decimal import Decimal, InvalidOperation
import re

from django.db import transaction
from django.utils.html import strip_tags
from django.utils.text import slugify
from openpyxl import load_workbook

from .models import Brand, Category, Product, ProductImage
from .serializers import build_unique_category_slug


DATA_START_ROW = 5
SUPPORTED_SHEET_NAME = 'Шаблон'

HEADER_ARTICLE = 'Артикул*'
HEADER_NAME = 'Название товара'
HEADER_PRICE = 'Цена, руб.*'
HEADER_MAIN_IMAGE = 'Ссылка на главное фото*'
HEADER_EXTRA_IMAGES = 'Ссылки на дополнительные фото'
HEADER_BRAND = 'Бренд*'
HEADER_DESCRIPTION = 'Аннотация'
HEADER_TYPE = 'Тип*'


def _normalize_header(value):
    return str(value or '').strip()


def _parse_price(value):
    raw = str(value or '').strip().replace(' ', '').replace(',', '.')
    if not raw:
        raise ValueError('Не заполнена цена')
    try:
        price = Decimal(raw)
    except InvalidOperation as exc:
        raise ValueError(f'Некорректная цена: {value}') from exc
    if price < 0:
        raise ValueError('Цена не может быть отрицательной')
    return price.quantize(Decimal('0.01'))


def _split_image_urls(raw_value):
    if not raw_value:
        return []
    parts = re.split(r'[\n;,]+', str(raw_value))
    return [part.strip() for part in parts if part and part.strip()]


def _normalize_description(value):
    raw = str(value or '').strip()
    if not raw:
        return ''
    raw = re.sub(r'<\s*br\s*/?>', '\n', raw, flags=re.IGNORECASE)
    raw = re.sub(r'</\s*(p|li|ul|ol)\s*>', '\n', raw, flags=re.IGNORECASE)
    text = strip_tags(raw)
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text.strip()


def _build_brand(name):
    brand_name = str(name or '').strip()
    if not brand_name:
        return None

    existing = Brand.objects.filter(name__iexact=brand_name).first()
    if existing:
        return existing

    base_slug = slugify(brand_name) or 'brand'
    slug = base_slug
    suffix = 2
    while Brand.objects.filter(slug=slug).exists():
        slug = f'{base_slug}-{suffix}'
        suffix += 1

    return Brand.objects.create(
        name=brand_name,
        slug=slug,
        is_active=True,
    )


def _resolve_category(type_name):
    value = str(type_name or '').strip()
    if not value:
        raise ValueError('Не заполнено поле "Тип"')

    existing = Category.objects.filter(name__iexact=value).order_by('id').first()
    if existing:
        return existing

    return Category.objects.create(
        name=value,
        slug=build_unique_category_slug(value),
        parent=None,
        is_active=True,
    )


def _build_product_slug(name, sku, instance=None):
    base_slug = slugify(name) or slugify(sku) or 'product'
    candidate = base_slug
    suffix = 2

    qs = Product.objects.all()
    if instance is not None:
        qs = qs.exclude(id=instance.id)

    while qs.filter(slug=candidate).exists():
        candidate = f'{base_slug}-{suffix}'
        suffix += 1
    return candidate


def _extract_sheet(workbook):
    if SUPPORTED_SHEET_NAME in workbook.sheetnames:
        return workbook[SUPPORTED_SHEET_NAME]
    return workbook.worksheets[0]


def _extract_columns(sheet):
    headers = {}
    for col_idx, cell in enumerate(sheet[2], start=1):
        header = _normalize_header(cell.value)
        if header:
            headers[header] = col_idx

    required_headers = [
        HEADER_ARTICLE,
        HEADER_NAME,
        HEADER_PRICE,
        HEADER_TYPE,
    ]
    missing = [header for header in required_headers if header not in headers]
    if missing:
        raise ValueError(f'В файле не хватает колонок: {", ".join(missing)}')
    return headers


def _cell(sheet, row_idx, headers, name):
    col_idx = headers.get(name)
    if not col_idx:
        return None
    return sheet.cell(row=row_idx, column=col_idx).value


def import_products_from_ozon_excel(uploaded_file, *, stock_default=0, is_active=True):
    workbook = load_workbook(uploaded_file, data_only=True)
    sheet = _extract_sheet(workbook)
    headers = _extract_columns(sheet)

    rows = []
    created = 0
    skipped = 0
    errors = 0

    for row_idx in range(DATA_START_ROW, sheet.max_row + 1):
        article = str(_cell(sheet, row_idx, headers, HEADER_ARTICLE) or '').strip()
        name = str(_cell(sheet, row_idx, headers, HEADER_NAME) or '').strip()
        raw_price = _cell(sheet, row_idx, headers, HEADER_PRICE)
        type_name = str(_cell(sheet, row_idx, headers, HEADER_TYPE) or '').strip()

        if not article and not name and raw_price in (None, ''):
            continue

        if not article:
            skipped += 1
            rows.append({
                'row_number': row_idx,
                'sku': '',
                'name': name,
                'action': 'skipped',
                'message': 'Пропущено: не заполнен артикул',
            })
            continue

        if not name:
            skipped += 1
            rows.append({
                'row_number': row_idx,
                'sku': article,
                'name': '',
                'action': 'skipped',
                'message': 'Пропущено: не заполнено название товара',
            })
            continue

        try:
            price = _parse_price(raw_price)
            brand = _build_brand(_cell(sheet, row_idx, headers, HEADER_BRAND))
            description = _normalize_description(_cell(sheet, row_idx, headers, HEADER_DESCRIPTION))
            category = _resolve_category(type_name)
            main_image = str(_cell(sheet, row_idx, headers, HEADER_MAIN_IMAGE) or '').strip()
            extra_images = _split_image_urls(_cell(sheet, row_idx, headers, HEADER_EXTRA_IMAGES))

            image_urls = []
            if main_image:
                image_urls.append(main_image)
            image_urls.extend([url for url in extra_images if url and url != main_image])

            existing_by_sku = Product.objects.filter(sku=article).first()
            existing_by_name = Product.objects.filter(name__iexact=name).first()
            if existing_by_sku or existing_by_name:
                duplicate_reason = []
                if existing_by_sku:
                    duplicate_reason.append(f'SKU #{existing_by_sku.id}')
                if existing_by_name and (not existing_by_sku or existing_by_name.id != existing_by_sku.id):
                    duplicate_reason.append(f'название #{existing_by_name.id}')

                skipped += 1
                rows.append({
                    'row_number': row_idx,
                    'sku': article,
                    'name': name,
                    'category_name': category.name,
                    'action': 'skipped',
                    'message': f'Пропущено: товар уже существует ({", ".join(duplicate_reason)})',
                })
                continue

            with transaction.atomic():
                product = Product(
                    sku=article,
                    category_id=category.id,
                )
                product.name = name
                product.slug = _build_product_slug(name, article)
                product.description = description
                product.category = category
                product.brand = brand
                product.price = price
                product.stock = stock_default
                product.is_active = is_active
                product.save()

                product.images.all().delete()
                if image_urls:
                    ProductImage.objects.bulk_create([
                        ProductImage(product=product, image_url=url, sort_order=index)
                        for index, url in enumerate(image_urls)
                    ])

            created += 1

            rows.append({
                'row_number': row_idx,
                'sku': article,
                'name': name,
                'category_name': category.name,
                'action': 'created',
                'product_id': product.id,
                'message': 'Сохранено',
            })
        except Exception as exc:
            errors += 1
            rows.append({
                'row_number': row_idx,
                'sku': article,
                'name': name,
                'category_name': type_name,
                'action': 'error',
                'message': str(exc),
            })

    return {
        'sheet_name': sheet.title,
        'summary': {
            'created': created,
            'skipped': skipped,
            'errors': errors,
            'processed': created + skipped + errors,
        },
        'rows': rows,
    }
