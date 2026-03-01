from datetime import date, datetime, time, timedelta
import csv
from io import BytesIO, StringIO

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.http import HttpResponse
from django.db.models import Q, CharField, Count, Sum, Avg, DecimalField, F, ExpressionWrapper, Exists, OuterRef
from django.db.models.functions import Cast
from django.db.models.functions import Coalesce, TruncDate, TruncWeek, TruncMonth, ExtractIsoWeekDay
from django.utils import timezone
from django.utils.dateparse import parse_date

from .models import Order, OrderItem
from .permissions import IsStaffOrWarehouseGroup
from .staff_serializers import StaffOrderStatusUpdateSerializer
from .serializers import serialize_order
from payments.models import Payment


STATUS_LABELS = {
    'new': 'Новые',
    'paid': 'Оплачены',
    'shipped': 'Отгружены',
    'completed': 'Доставлены',
    'canceled': 'Отменены',
}

DELIVERY_LABELS = {
    'store_pickup': 'Самовывоз',
    'courier': 'Курьер',
    'pvz': 'ПВЗ',
}

PAYMENT_STATUS_LABELS = {
    Payment.Status.PENDING: 'В ожидании',
    Payment.Status.WAITING_FOR_CAPTURE: 'В обработке',
    Payment.Status.SUCCEEDED: 'Успешно',
    Payment.Status.CANCELED: 'Отменен',
    Payment.Status.FAILED: 'Ошибка',
}

WEEKDAY_LABELS = {
    1: 'Пн',
    2: 'Вт',
    3: 'Ср',
    4: 'Чт',
    5: 'Пт',
    6: 'Сб',
    7: 'Вс',
}


def _delivery_label(value):
    return DELIVERY_LABELS.get(value, value or '')


def _status_label(value):
    return STATUS_LABELS.get(value, value or '')


def _payment_status_label(value):
    return PAYMENT_STATUS_LABELS.get(value, value or '')


def _pick_payment(order):
    payments = list(order.payments.all())
    if not payments:
        return None
    succeeded = next((payment for payment in payments if payment.status == Payment.Status.SUCCEEDED), None)
    return succeeded or payments[0]


def _build_export_order_rows(orders_qs):
    rows = []
    export_orders = (
        orders_qs
        .prefetch_related('items', 'payments')
        .order_by('-created_at', '-id')
    )

    for order in export_orders:
        payment = _pick_payment(order)
        items = list(order.items.all())
        items_count = sum(item.quantity for item in items)
        items_summary = '; '.join(
            f"{item.product_name_snapshot} x{item.quantity}"
            for item in items
            if item.product_name_snapshot
        )

        rows.append({
            'id': order.id,
            'status': order.status,
            'status_label': _status_label(order.status),
            'customer_name': order.customer_name or '',
            'customer_phone': order.customer_phone or '',
            'customer_email': order.customer_email or '',
            'delivery_type': order.delivery_type,
            'delivery_type_label': _delivery_label(order.delivery_type),
            'delivery_city': order.delivery_city or '',
            'delivery_address_text': order.delivery_address_text or '',
            'delivery_service': order.delivery_service or '',
            'total_amount': order.total_amount or 0,
            'items_count': items_count,
            'items_summary': items_summary,
            'payment_status': payment.status if payment else '',
            'payment_status_label': _payment_status_label(payment.status) if payment else '',
            'payment_amount': payment.amount_value if payment else '',
            'payment_paid_at': payment.paid_at if payment else None,
            'created_at': order.created_at,
            'updated_at': order.updated_at,
        })

    return rows


def _build_analytics_payload(date_from, date_to, start, end):
    orders_qs = (
        Order.objects
        .filter(created_at__gte=start, created_at__lte=end)
    )
    payments_qs = Payment.objects.filter(created_at__gte=start, created_at__lte=end)
    paid_payments_qs = Payment.objects.filter(
        status=Payment.Status.SUCCEEDED,
        paid_at__gte=start,
        paid_at__lte=end,
    )

    overview = orders_qs.aggregate(
        orders_total=Count('id'),
        gross_revenue=Coalesce(Sum('total_amount'), 0, output_field=DecimalField(max_digits=12, decimal_places=2)),
        average_order=Coalesce(Avg('total_amount'), 0, output_field=DecimalField(max_digits=12, decimal_places=2)),
        pending_payment=Count('id', filter=Q(status='new')),
        pending_assembly=Count('id', filter=Q(status='paid')),
        shipped_total=Count('id', filter=Q(status='shipped')),
        completed_total=Count('id', filter=Q(status='completed')),
        canceled_total=Count('id', filter=Q(status='canceled')),
    )

    total_items = OrderItem.objects.filter(order__in=orders_qs).aggregate(
        total_items=Coalesce(Sum('quantity'), 0)
    )['total_items']

    paid_revenue = paid_payments_qs.aggregate(
        paid_revenue=Coalesce(Sum('amount_value'), 0, output_field=DecimalField(max_digits=12, decimal_places=2))
    )['paid_revenue']
    delivered_revenue = orders_qs.filter(status='completed').aggregate(
        delivered_revenue=Coalesce(Sum('total_amount'), 0, output_field=DecimalField(max_digits=12, decimal_places=2))
    )['delivered_revenue']

    payments_total = payments_qs.count()
    succeeded_payments = payments_qs.filter(status=Payment.Status.SUCCEEDED).count()

    orders_total = overview['orders_total'] or 0
    canceled_total = overview['canceled_total'] or 0
    completed_total = overview['completed_total'] or 0
    shipped_total = overview['shipped_total'] or 0
    completion_rate = round((completed_total / orders_total) * 100, 1) if orders_total else 0
    shipping_rate = round((shipped_total / orders_total) * 100, 1) if orders_total else 0
    cancellation_rate = round((canceled_total / orders_total) * 100, 1) if orders_total else 0
    payment_success_rate = round((succeeded_payments / payments_total) * 100, 1) if payments_total else 0

    granularity, timeline = _build_timeline_series(orders_qs, paid_payments_qs, date_from, date_to)

    status_breakdown = [
        {
            'key': key,
            'label': STATUS_LABELS.get(key, key),
            'count': orders_qs.filter(status=key).count(),
            'share': round((orders_qs.filter(status=key).count() / orders_total) * 100, 1) if orders_total else 0,
        }
        for key in ['new', 'paid', 'shipped', 'completed', 'canceled']
    ]

    delivery_breakdown = _build_delivery_breakdown(orders_qs)
    payment_breakdown = _build_payment_breakdown(payments_qs)

    top_products = list(
        orders_qs
        .exclude(status='canceled')
        .values('items__product_id', 'items__product_name_snapshot')
        .annotate(
            units=Coalesce(Sum('items__quantity'), 0),
            revenue=Coalesce(
                Sum(
                    ExpressionWrapper(
                        F('items__price_snapshot') * F('items__quantity'),
                        output_field=DecimalField(max_digits=12, decimal_places=2),
                    )
                ),
                0,
                output_field=DecimalField(max_digits=12, decimal_places=2),
            ),
            orders_count=Count('id', distinct=True),
        )
        .order_by('-units', '-revenue')[:8]
    )

    recent_orders = [
        {
            'id': order.id,
            'status': order.status,
            'customer_name': order.customer_name,
            'delivery_type': order.delivery_type,
            'total_amount': order.total_amount,
            'created_at': order.created_at,
            'items_count': sum(item.quantity for item in order.items.all()),
        }
        for order in orders_qs.prefetch_related('items').order_by('-created_at')[:8]
    ]

    attention = [
        {
            'key': 'gross_revenue',
            'label': 'Заказано на сумму',
            'value': overview['gross_revenue'] or 0,
            'format': 'money',
            'tone': 'neutral',
        },
        {
            'key': 'paid_revenue',
            'label': 'Оплачено на сумму',
            'value': paid_revenue or 0,
            'format': 'money',
            'tone': 'info',
        },
        {
            'key': 'payment_success_rate',
            'label': 'Успешные оплаты',
            'value': payment_success_rate,
            'format': 'percent',
            'tone': 'neutral',
        },
        {
            'key': 'cancellation_rate',
            'label': 'Процент отмен',
            'value': cancellation_rate,
            'format': 'percent',
            'tone': 'danger' if cancellation_rate >= 15 else 'neutral',
        },
    ]

    return {
        'period': {
            'date_from': date_from.isoformat(),
            'date_to': date_to.isoformat(),
            'days': (date_to - date_from).days + 1,
            'granularity': granularity,
        },
        'overview': {
            'orders_total': orders_total,
            'gross_revenue': overview['gross_revenue'] or 0,
            'paid_revenue': paid_revenue or 0,
            'delivered_revenue': delivered_revenue or 0,
            'average_order': overview['average_order'] or 0,
            'avg_items_per_order': round((float(total_items or 0) / orders_total), 2) if orders_total else 0,
            'pending_payment': overview['pending_payment'] or 0,
            'pending_assembly': overview['pending_assembly'] or 0,
            'shipped_total': shipped_total,
            'completed_total': completed_total,
            'canceled_total': canceled_total,
            'completion_rate': completion_rate,
            'shipping_rate': shipping_rate,
            'cancellation_rate': cancellation_rate,
            'payments_total': payments_total,
            'payment_success_rate': payment_success_rate,
        },
        'status_breakdown': status_breakdown,
        'delivery_breakdown': delivery_breakdown,
        'payment_breakdown': payment_breakdown,
        'timeline': timeline,
        'funnel': _build_funnel(orders_qs),
        'weekday_breakdown': _build_weekday_breakdown(orders_qs),
        'city_breakdown': _build_city_breakdown(orders_qs),
        'top_products': [
            {
                'product_id': row['items__product_id'],
                'name': row['items__product_name_snapshot'],
                'units': row['units'],
                'revenue': row['revenue'],
                'orders_count': row['orders_count'],
            }
            for row in top_products
            if row['items__product_name_snapshot']
        ],
        'recent_orders': recent_orders,
        'attention': attention,
        'export_orders': _build_export_order_rows(orders_qs),
    }


def _build_csv_response(payload):
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'Номер заказа',
        'Статус',
        'Клиент',
        'Телефон',
        'Email',
        'Тип доставки',
        'Город',
        'Адрес',
        'Служба доставки',
        'Сумма заказа',
        'Кол-во товаров',
        'Состав',
        'Статус оплаты',
        'Сумма платежа',
        'Оплачен',
        'Создан',
        'Обновлен',
    ])
    for row in payload['export_orders']:
        writer.writerow([
            row['id'],
            row['status_label'],
            row['customer_name'],
            row['customer_phone'],
            row['customer_email'],
            row['delivery_type_label'],
            row['delivery_city'],
            row['delivery_address_text'],
            row['delivery_service'],
            row['total_amount'],
            row['items_count'],
            row['items_summary'],
            row['payment_status_label'],
            row['payment_amount'],
            row['payment_paid_at'].strftime('%Y-%m-%d %H:%M:%S') if row['payment_paid_at'] else '',
            row['created_at'].strftime('%Y-%m-%d %H:%M:%S'),
            row['updated_at'].strftime('%Y-%m-%d %H:%M:%S'),
        ])

    response = HttpResponse(output.getvalue(), content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = (
        f'attachment; filename="analytics_{payload["period"]["date_from"]}_{payload["period"]["date_to"]}.csv"'
    )
    return response


def _build_xlsx_response(payload):
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError('Для экспорта Excel нужен пакет openpyxl') from exc

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = 'Сводка'
    summary_ws.append(['Период', f'{payload["period"]["date_from"]} - {payload["period"]["date_to"]}'])
    summary_ws.append(['Создано заказов', payload['overview']['orders_total']])
    summary_ws.append(['Заказано на сумму', float(payload['overview']['gross_revenue'])])
    summary_ws.append(['Оплачено на сумму', float(payload['overview']['paid_revenue'])])
    summary_ws.append(['Доставленная выручка', float(payload['overview']['delivered_revenue'])])
    summary_ws.append(['Средний чек', float(payload['overview']['average_order'])])
    summary_ws.append(['Среднее число товаров', payload['overview']['avg_items_per_order']])
    summary_ws.append(['Ожидают оплаты', payload['overview']['pending_payment']])
    summary_ws.append(['К сборке', payload['overview']['pending_assembly']])
    summary_ws.append(['В пути', payload['overview']['shipped_total']])
    summary_ws.append(['Доставлены', payload['overview']['completed_total']])
    summary_ws.append(['Отменены', payload['overview']['canceled_total']])

    orders_ws = wb.create_sheet('Заказы')
    orders_ws.append([
        'Номер заказа', 'Статус', 'Клиент', 'Телефон', 'Email', 'Тип доставки',
        'Город', 'Адрес', 'Служба доставки', 'Сумма заказа', 'Кол-во товаров',
        'Состав', 'Статус оплаты', 'Сумма платежа', 'Оплачен', 'Создан', 'Обновлен',
    ])
    for row in payload['export_orders']:
        orders_ws.append([
            row['id'],
            row['status_label'],
            row['customer_name'],
            row['customer_phone'],
            row['customer_email'],
            row['delivery_type_label'],
            row['delivery_city'],
            row['delivery_address_text'],
            row['delivery_service'],
            float(row['total_amount'] or 0),
            row['items_count'],
            row['items_summary'],
            row['payment_status_label'],
            float(row['payment_amount'] or 0) if row['payment_amount'] != '' else '',
            row['payment_paid_at'].strftime('%Y-%m-%d %H:%M:%S') if row['payment_paid_at'] else '',
            row['created_at'].strftime('%Y-%m-%d %H:%M:%S'),
            row['updated_at'].strftime('%Y-%m-%d %H:%M:%S'),
        ])

    products_ws = wb.create_sheet('Топ товаров')
    products_ws.append(['Товар', 'ID товара', 'Штук', 'Выручка', 'Заказов'])
    for row in payload['top_products']:
        products_ws.append([
            row['name'],
            row['product_id'],
            row['units'],
            float(row['revenue'] or 0),
            row['orders_count'],
        ])

    timeline_ws = wb.create_sheet('Динамика')
    timeline_ws.append([
        'Начало периода', 'Конец периода', 'Заказов', 'Заказано на сумму',
        'Оплачено на сумму', 'Доставленная выручка', 'В пути', 'Доставлено', 'Отменено',
    ])
    for row in payload['timeline']:
        timeline_ws.append([
            row['bucket_start'],
            row['bucket_end'],
            row['orders'],
            float(row['revenue'] or 0),
            float(row['paid_revenue'] or 0),
            float(row['delivered_revenue'] or 0),
            row['shipped_orders'],
            row['completed_orders'],
            row['canceled_orders'],
        ])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = (
        f'attachment; filename="analytics_{payload["period"]["date_from"]}_{payload["period"]["date_to"]}.xlsx"'
    )
    return response


def _parse_period(request):
    today = timezone.localdate()
    default_from = today - timedelta(days=29)

    date_from = parse_date(request.query_params.get('date_from') or '') or default_from
    date_to = parse_date(request.query_params.get('date_to') or '') or today

    if date_from > date_to:
        date_from, date_to = date_to, date_from

    tz = timezone.get_current_timezone()
    start = timezone.make_aware(datetime.combine(date_from, time.min), tz)
    end = timezone.make_aware(datetime.combine(date_to, time.max), tz)

    return date_from, date_to, start, end


def _resolve_granularity(days):
    if days <= 45:
        return 'day'
    if days <= 180:
        return 'week'
    return 'month'


def _month_start(value):
    return value.replace(day=1)


def _month_step(value):
    if value.month == 12:
        return date(value.year + 1, 1, 1)
    return date(value.year, value.month + 1, 1)


def _period_label(bucket_start, bucket_end, granularity):
    if granularity == 'day':
        return bucket_start.strftime('%d %b')
    if granularity == 'week':
        if bucket_start == bucket_end:
            return bucket_start.strftime('%d %b')
        return f"{bucket_start.strftime('%d %b')} - {bucket_end.strftime('%d %b')}"
    return bucket_start.strftime('%b %Y')


def _bucket_starts(date_from, date_to, granularity):
    if granularity == 'day':
        current = date_from
        while current <= date_to:
            yield current
            current += timedelta(days=1)
        return

    if granularity == 'week':
        current = date_from - timedelta(days=date_from.weekday())
        while current <= date_to:
            yield current
            current += timedelta(days=7)
        return

    current = _month_start(date_from)
    while current <= date_to:
        yield current
        current = _month_step(current)


def _bucket_end(bucket_start, date_to, granularity):
    if granularity == 'day':
        return min(bucket_start, date_to)
    if granularity == 'week':
        return min(bucket_start + timedelta(days=6), date_to)
    return min(_month_step(bucket_start) - timedelta(days=1), date_to)


def _normalize_bucket_value(value):
    if hasattr(value, 'date'):
        try:
            return value.date()
        except TypeError:
            pass
    return value


def _build_timeline_series(orders_qs, paid_payments_qs, date_from, date_to):
    days = (date_to - date_from).days + 1
    granularity = _resolve_granularity(days)

    if granularity == 'day':
        trunc = TruncDate('created_at')
        payment_trunc = TruncDate('paid_at')
    elif granularity == 'week':
        trunc = TruncWeek('created_at')
        payment_trunc = TruncWeek('paid_at')
    else:
        trunc = TruncMonth('created_at')
        payment_trunc = TruncMonth('paid_at')

    raw_series = {
        _normalize_bucket_value(row['bucket']): row
        for row in (
            orders_qs
            .annotate(bucket=trunc)
            .values('bucket')
            .annotate(
                orders=Count('id'),
                revenue=Coalesce(Sum('total_amount'), 0, output_field=DecimalField(max_digits=12, decimal_places=2)),
                paid_revenue=Coalesce(
                    Sum('total_amount', filter=Q(status__in=['paid', 'shipped', 'completed'])),
                    0,
                    output_field=DecimalField(max_digits=12, decimal_places=2),
                ),
                delivered_revenue=Coalesce(
                    Sum('total_amount', filter=Q(status='completed')),
                    0,
                    output_field=DecimalField(max_digits=12, decimal_places=2),
                ),
                paid_orders=Count('id', filter=Q(status='paid')),
                shipped_orders=Count('id', filter=Q(status='shipped')),
                completed_orders=Count('id', filter=Q(status='completed')),
                canceled_orders=Count('id', filter=Q(status='canceled')),
            )
            .order_by('bucket')
        )
    }

    paid_series = {
        _normalize_bucket_value(row['bucket']): row['paid_revenue']
        for row in (
            paid_payments_qs
            .annotate(bucket=payment_trunc)
            .values('bucket')
            .annotate(
                paid_revenue=Coalesce(Sum('amount_value'), 0, output_field=DecimalField(max_digits=12, decimal_places=2))
            )
            .order_by('bucket')
        )
    }

    series = []
    for bucket_start in _bucket_starts(date_from, date_to, granularity):
        bucket_end = _bucket_end(bucket_start, date_to, granularity)
        row = raw_series.get(bucket_start, {})
        orders = row.get('orders', 0) or 0
        shipped_orders = row.get('shipped_orders', 0) or 0
        completed_orders = row.get('completed_orders', 0) or 0
        delivery_flow_orders = shipped_orders + completed_orders
        completion_rate = round((completed_orders / orders) * 100, 1) if orders else 0
        shipped_rate = round((shipped_orders / orders) * 100, 1) if orders else 0
        delivery_flow_rate = round((delivery_flow_orders / orders) * 100, 1) if orders else 0
        delivered_share_in_flow = round((completed_orders / delivery_flow_orders) * 100, 1) if delivery_flow_orders else 0

        series.append({
            'bucket_start': bucket_start.isoformat(),
            'bucket_end': bucket_end.isoformat(),
            'label': _period_label(bucket_start, bucket_end, granularity),
            'orders': orders,
            'revenue': row.get('revenue', 0) or 0,
            'paid_revenue': paid_series.get(bucket_start, 0) or 0,
            'delivered_revenue': row.get('delivered_revenue', 0) or 0,
            'paid_orders': row.get('paid_orders', 0) or 0,
            'shipped_orders': shipped_orders,
            'completed_orders': completed_orders,
            'canceled_orders': row.get('canceled_orders', 0) or 0,
            'delivery_flow_orders': delivery_flow_orders,
            'completion_rate': completion_rate,
            'shipped_rate': shipped_rate,
            'delivery_flow_rate': delivery_flow_rate,
            'delivered_share_in_flow': delivered_share_in_flow,
        })

    return granularity, series


def _build_funnel(orders_qs):
    total_created = orders_qs.count()
    paid = orders_qs.filter(status__in=['paid', 'shipped', 'completed']).count()
    shipped = orders_qs.filter(status__in=['shipped', 'completed']).count()
    delivered = orders_qs.filter(status='completed').count()
    return [
        {'key': 'new', 'label': 'Созданы', 'value': total_created},
        {'key': 'paid', 'label': 'Оплачены', 'value': paid},
        {'key': 'shipped', 'label': 'Отгружены', 'value': shipped},
        {'key': 'completed', 'label': 'Доставлены', 'value': delivered},
    ]


def _build_weekday_breakdown(orders_qs):
    raw = {
        row['weekday']: row['count']
        for row in (
            orders_qs
            .annotate(weekday=ExtractIsoWeekDay('created_at'))
            .values('weekday')
            .annotate(count=Count('id'))
            .order_by('weekday')
        )
    }
    return [
        {
            'weekday': index,
            'label': WEEKDAY_LABELS[index],
            'count': raw.get(index, 0),
        }
        for index in range(1, 8)
    ]


def _build_city_breakdown(orders_qs):
    rows = (
        orders_qs
        .exclude(delivery_city__isnull=True)
        .exclude(delivery_city__exact='')
        .values('delivery_city')
        .annotate(
            orders=Count('id'),
            revenue=Coalesce(Sum('total_amount'), 0, output_field=DecimalField(max_digits=12, decimal_places=2)),
        )
        .order_by('-orders', '-revenue')[:6]
    )
    return [
        {
            'city': row['delivery_city'],
            'orders': row['orders'],
            'revenue': row['revenue'],
        }
        for row in rows
    ]


def _build_delivery_breakdown(orders_qs):
    total = orders_qs.count() or 1
    items = []
    for key in ['store_pickup', 'courier', 'pvz']:
        count = orders_qs.filter(delivery_type=key).count()
        items.append({
            'key': key,
            'label': DELIVERY_LABELS.get(key, key),
            'count': count,
            'share': round((count / total) * 100, 1) if total else 0,
        })
    return items


def _build_payment_breakdown(payments_qs):
    total = payments_qs.count() or 1
    items = []
    for key in [
        Payment.Status.PENDING,
        Payment.Status.WAITING_FOR_CAPTURE,
        Payment.Status.SUCCEEDED,
        Payment.Status.CANCELED,
        Payment.Status.FAILED,
    ]:
        count = payments_qs.filter(status=key).count()
        items.append({
            'key': key,
            'label': PAYMENT_STATUS_LABELS.get(key, key),
            'count': count,
            'share': round((count / total) * 100, 1) if total else 0,
        })
    return items


class StaffAnalyticsView(APIView):
    permission_classes = [IsStaffOrWarehouseGroup]

    def get(self, request):
        date_from, date_to, start, end = _parse_period(request)
        payload = _build_analytics_payload(date_from, date_to, start, end)
        payload.pop('export_orders', None)
        return Response(payload, status=status.HTTP_200_OK)


class StaffAnalyticsExportView(APIView):
    permission_classes = [IsStaffOrWarehouseGroup]

    def get(self, request):
        export_format = (
            request.query_params.get('export_format')
            or request.query_params.get('format')
            or 'csv'
        ).lower()
        date_from, date_to, start, end = _parse_period(request)
        payload = _build_analytics_payload(date_from, date_to, start, end)

        if export_format == 'csv':
            return _build_csv_response(payload)
        if export_format in {'xlsx', 'excel'}:
            try:
                return _build_xlsx_response(payload)
            except RuntimeError as exc:
                return Response({'detail': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response({'detail': 'Поддерживаются только форматы csv и xlsx'}, status=status.HTTP_400_BAD_REQUEST)


class StaffOrdersListView(APIView):
    permission_classes = [IsStaffOrWarehouseGroup]

    @staticmethod
    def base_queryset():
        return (
            Order.objects
            .prefetch_related('items', 'payments')
            .order_by('-id')
        )

    def get(self, request):
        status_q = request.query_params.get('status')
        delivery_type = request.query_params.get('delivery_type')
        q = request.query_params.get('q')

        qs = self.base_queryset()

        if status_q:
            qs = qs.filter(status=status_q)
        if delivery_type:
            qs = qs.filter(delivery_type=delivery_type)
        if q:
            qs = qs.annotate(id_text=Cast('id', CharField())).filter(
                Q(id_text__icontains=q) |
                Q(customer_name__icontains=q) |
                Q(customer_phone__icontains=q)
            )

        return Response({'results': [serialize_order(order, include_customer=True) for order in qs]}, status=status.HTTP_200_OK)


class StaffOrderDetailView(APIView):
    permission_classes = [IsStaffOrWarehouseGroup]

    def get(self, request, order_id):
        try:
            order = StaffOrdersListView.base_queryset().get(id=order_id)
        except Order.DoesNotExist:
            return Response({'detail': 'Заказ не найден'}, status=status.HTTP_404_NOT_FOUND)
        return Response(serialize_order(order, include_customer=True), status=status.HTTP_200_OK)


class StaffOrderStatusUpdateView(APIView):
    permission_classes = [IsStaffOrWarehouseGroup]

    def patch(self, request, order_id):
        try:
            order = StaffOrdersListView.base_queryset().get(id=order_id)
        except Order.DoesNotExist:
            return Response({'detail': 'Заказ не найден'}, status=status.HTTP_404_NOT_FOUND)

        serializer = StaffOrderStatusUpdateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        new_status = serializer.validated_data['status']
        order.status = new_status
        order.save(update_fields=['status', 'updated_at'])
        order.refresh_from_db()

        return Response(serialize_order(order, include_customer=True), status=status.HTTP_200_OK)

    def post(self, request, order_id):
        return self.patch(request, order_id)
